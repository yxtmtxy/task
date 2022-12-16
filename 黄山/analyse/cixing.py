from ltp import LTP
import pandas as pd
from openpyxl import load_workbook
import openpyxl
ltp = LTP()
usewords = openpyxl.load_workbook(r"D:\python\程序\黄山\analyse\result1.xlsx")
sheet1 = usewords['Sheet1']
line=2
i1=0
re={}
result={}
j=1

while line<958:
    cell = sheet1.cell(row=line, column=2)
    print(cell.value)
    re[i1] = ltp.pipeline(cell.value, tasks=["cws","pos"])
    print(re[i1].pos)
    sheet1.cell(line, 4).value = str(re[i1].pos)
    line+=1
    i1+=1
usewords.save("result1.xlsx")