from ltp import LTP
import pandas as pd
from openpyxl import load_workbook
import openpyxl

def stopwordslist():#停用词词典引入
    stopwords = [line.strip() for line in open(r"/analyse/cn_stopwords.txt", encoding='UTF-8').readlines()]
    return stopwords

usewords = openpyxl.load_workbook(r"/analyse/sum1.xlsx")#导入爬虫文件
sheet1 = usewords['1']
sheet2 = usewords['2']
sheet3 = usewords['3']
sheet4 = usewords['4']#sum文件中表格sheet

ltp = LTP()

stopwords = stopwordslist()
outword={}
output={}
j = 0#总分词数
def depart (sheet,l_max):#实现分词函数
    line = 1
    global j
    while (line < l_max):
        r_first = j
        cell = sheet.cell(row=line, column=2)
        print(cell.value)
        if(cell.value is not None):
           output[line] = ltp.pipeline(cell.value, tasks=["cws", "pos", "ner", "srl", "dep", "sdp"])
           for i in output[line].cws:
               if i not in stopwords:
                  outword[j] = i
                  print(outword[j])
                  j += 1
        for a in range(j - r_first):
           sheet.cell(line, 3 + a).value = outword[r_first + a]
        line += 1
    return

depart(sheet1,687)
depart(sheet2,1075)
depart(sheet3,75)
depart(sheet4,1427)
usewords.save("sum1.xlsx")

sum = pd.Series(outword)#转为词典
countDict = dict(sum.value_counts().iloc[:1000])#取前1000条分词结果
print(countDict)

key = list(countDict.keys())
value = list(countDict.values())
result = pd.DataFrame()
result["词向量"] = key
result["词频"] = value
# 写入excel
result.to_excel('result1.xlsx')
