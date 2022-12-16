import openpyxl
import numpy as np


def read_file(file):
    text = [line.strip() for line in open(file, encoding='GBK').readlines()]
    return text


def weighted_value(request):
    result_dict = []
    if request == "one":
        result_dict = read_file(r"D:\python\程序\天柱山\emotion\正面情感词语（中文）.txt")
    elif request == "two":
        result_dict = read_file(r"D:\python\程序\天柱山\emotion\正面评价词语（中文）.txt")
    elif request == "three":
        result_dict = read_file(r"D:\python\程序\天柱山\emotion\程度级别词语（中文）.txt")
    elif request == "four":
        result_dict = read_file(r"D:\python\程序\天柱山\emotion\负面情感词语（中文）.txt")
    elif request == "five":
        result_dict = read_file(r"D:\python\程序\天柱山\emotion\负面评价词语（中文）.txt")
    elif request == "six":
        result_dict = read_file(r"D:\python\程序\天柱山\emotion\主张词语（中文）.txt")
    else:
        pass
    return result_dict


po_emotion = weighted_value('one')  # 正面情感
po_comment = weighted_value('two')  # 正面评价
extent = weighted_value('three')  # 程度副词
ne_emotion = weighted_value('four')  # 负面情感
ne_comment = weighted_value('five')  # 负面评价
perception = weighted_value('six')  # 主张副词


# 程度副词比重（待改进）
def match_adverb(word, value):
    # “极其|extreme / 最|most”
    if word in extent[3:71]:
        value *= 2.6
    # “很|very”
    elif word in extent[74:115]:
        value *= 2.3
    # “较|more”
    elif word in extent[118:154]:
        value *= 1.9
    # “稍|-ish”
    elif word in extent[157:185]:
        value *= 1.6
    # “欠|insufficiently”
    elif word in extent[188:199]:
        value *= 1.3
    # “超|over”
    elif word in extent[202:231]:
        value *= 0.5
    return value


usewords = openpyxl.load_workbook(r"D:\python\程序\天柱山\analyse\sum.xlsx")
emotion = openpyxl.load_workbook(r"D:\python\程序\天柱山\emotion\sum_em.xlsx")

sheet1 = usewords['1']
sheet2 = usewords['2']
sheet3 = usewords['3']
sheet4 = usewords['4']
sheet0 = emotion['1']


def depart(sheet, l_max):
    row1 = 1
    a = np.zeros(300)
    word = a.astype(np.str)

    value = np.zeros(l_max + 1)
    while row1 <= l_max:
        po_value = 0
        ne_value = 0
        i = 0
        s = 0
        col = 3
        cell = sheet.cell(row=row1, column=col)
        while cell.value is not None:
            word[i] = cell.value
            if word[i] in po_emotion[2:]:
                ps = 1
                # 每个正面情感词语得分2
                for w in word[s:i]:
                    ps = match_adverb(w, ps)
                po_value += 2 * ps
                s = i + 1
            elif word[i] in po_comment[2:]:
                ps = 1
                # 每个正面评价词语得分1
                for w in word[s:i]:
                    ps = match_adverb(w, ps)
                po_value += 1 * ps
                s = i + 1
            elif word[i] in ne_emotion[2:]:
                ps = 1
                # 每个负面情感词语得分2
                for w in word[s:i]:
                    ps = match_adverb(w, ps)
                ne_value -= 2 * ps
                s = i + 1
            elif word[i] in ne_comment[2:]:
                ps = 1
                # 每个负面情感词语得分1
                for w in word[s:i]:
                    ps = match_adverb(w, ps)
                ne_value -= 1 * ps
                s = i + 1

            col += 1
            i += 1
            cell = sheet.cell(row=row1, column=col)
        sheet0.cell(row1 + 1841, 3).value = str(po_value)  # 数据存贮进单元格 row+x x为存贮开始行数，3为列数
        sheet0.cell(row1 + 1841, 4).value = str(ne_value)
        value[row1] = po_value + ne_value #情感得分为正面得分减负面得分
        sheet0.cell(row1 + 1841, 2).value = str(value[row1])
        print(row1, value[row1], po_value, ne_value)
        row1 += 1
    return ()


depart(sheet4, 1426)
emotion.save("sum_em1.xlsx")

